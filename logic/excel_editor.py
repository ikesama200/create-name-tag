import shutil
import os
import sys
import re
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.pagebreak import RowBreak
from datetime import datetime
from copy import copy
import logging

from models.name_tag import NameTag

def run_export(data):
  logging.info("出力処理開始")
  if not data.template_path:
    raise ValueError("テンプレートが選択されていません")

  template = data.template_path
  value = data.value

  if not template:
      return

  os.makedirs("output", exist_ok=True)
  out_file = f"output/result_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

  shutil.copy(template, out_file)

  wb = load_workbook(out_file)
  ws = wb.active
  ws["A1"] = value
  wb.save(out_file)
  logging.info("出力処理終了")

# -------------------------
# Excelテンプレートパス取得
# -------------------------
def resource_path(relative_path):
  if hasattr(sys, "_MEIPASS"):
      return os.path.join(sys._MEIPASS, relative_path)
  return os.path.join(os.path.abspath("."), relative_path)

# -------------------------
# 名札情報から空のデータを削除して再作成する処理
# -------------------------
def recreate_name_tag(name_tag):
  # 再作成する名札データを作成
  new_name_tag = {}
  idx = 0
  for i, row in name_tag.items():
    # 行番号または行番号以外が空の場合はスキップする
    if (row.value == "") or (row.name == "" and row.nameKana == "" and row.itemA == "" and row.itemB == ""):
      continue
    # 再作成用の名札データに追加する(行番号は新しく採番)
    new_name_tag[idx] = NameTag(
        value=idx,
        name=row.name,
        nameKana=row.nameKana,
        itemA=row.itemA,
        itemB=row.itemB
    )
    # インデックスを進める
    idx += 1
  return new_name_tag

# -------------------------
# Excel書き込み処理の実行
# -------------------------
def load_excel_file(name_tag):
  # Excelテンプレートのパスを取得
  template_path = resource_path("templates/template.xlsx")
  # 出力先のパスを生成
  os.makedirs("output", exist_ok=True)
  output_path = f"output/result_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
  # 名札情報から空のデータを削除
  new_name_tag = recreate_name_tag(name_tag)
  # Excelへの書き込み処理を実行
  write_to_excel(new_name_tag, template_path, output_path)
  # 名札表示用のシートの調整処理を実行
  write_to_nametag_sheet(new_name_tag, output_path)

# -------------------------
# Excelへの書き込み処理
# -------------------------
def write_to_excel(name_tag, template_path, output_path):
  # テンプレートを呼び出し
  wb = load_workbook(template_path)
  # データ書き込み用のシートを選択
  ws = wb["List"]
  # データ書き込み
  for i, row in name_tag.items():
    # 書き込み行をセット(2行目からスタート)
    execl_row = i + 2
    # A列 空行
    # B列 行番号
    ws.cell(row=execl_row, column=2).value = row.value + 1
    # C列 名前
    ws.cell(row=execl_row, column=3).value = row.name
    # D列 仮名
    ws.cell(row=execl_row, column=4).value = row.nameKana
    # E列 項目A
    ws.cell(row=execl_row, column=5).value = row.itemA
    # F列 項目B
    ws.cell(row=execl_row, column=6).value = row.itemB
  # 書き込んだExcelを出力先のファイル名で保存
  wb.save(output_path)

# -------------------------
# 表示用シートへの書き込み処理
# -------------------------
def write_to_nametag_sheet(name_tag, output_path):
  # テンプレートを呼び出し
  wb = load_workbook(output_path)
  # データ書き込み用のシートを選択
  ws = wb["NameTag"]
  for i, row in name_tag.items():
    row_base = 2 + (i // 2) * 3
    # 現在のデータが2番目以降の場合はセルの書式をコピーする
    if i >= 2:
        copy_output_block(ws, 2, row_base)
  # 名札シートに改ページ設定を追加
  set_page_break(ws, 15)
  # 書き込んだExcelを出力先のファイル名で保存
  wb.save(output_path)

# -------------------------
# 出力用のセルを書式ごとコピー
# -------------------------
def copy_output_block(ws, src_row, dest_row):
  row_offset = dest_row - src_row
  for r in range(3):
    for c in range(1, 5):
        src_cell = ws.cell(row=src_row + r, column=c)
        dest_cell = ws.cell(row=dest_row + r, column=c)

        # ★ 追加：MergedCellはスキップ
        if isinstance(src_cell, MergedCell) or isinstance(dest_cell, MergedCell):
            continue

        # セルの値をコピー
        if isinstance(src_cell.value, str) and src_cell.value.startswith("="):
          dest_cell.value = shift_formula(src_cell.value, row_offset)
        else:
          dest_cell.value = src_cell.value
        
        # セルの書式をコピー
        if src_cell.has_style:
          dest_cell.font = copy(src_cell.font)
          dest_cell.border = copy(src_cell.border)
          dest_cell.fill = copy(src_cell.fill)
          dest_cell.number_format = src_cell.number_format
          dest_cell.alignment = copy(src_cell.alignment)
    # 行の高さを取得
    src_dim = ws.row_dimensions[src_row + r]
    dest_dim = ws.row_dimensions[dest_row + r]
    # コピー元のセルの高さを確認(デフォルトがNone)
    if src_dim.height is not None:
        # コピー先のセルに行の高さをコピーする
        dest_dim.height = src_dim.height

  for merged in list(ws.merged_cells.ranges):
    min_col, min_row, max_col, max_row = range_boundaries(str(merged))
    # コピー元ブロック内だけ対象
    if src_row <= min_row <= src_row + 2:
      new_min_row = min_row + row_offset
      new_max_row = max_row + row_offset

      new_range = f"{ws.cell(new_min_row, min_col).coordinate}:{ws.cell(new_max_row, max_col).coordinate}"

      ws.merge_cells(new_range)

# -------------------------
# セルの入力参照先の置換処理
# -------------------------
def shift_formula(formula, row_offset):
  def repl(match):
    col = match.group(1)
    row = int(match.group(2))
    return f"{col}{row + ((row_offset // 3) * 2)}"

  return re.sub(r'([A-Z]+)(\d+)', repl, formula)
# -------------------------
# シートへの改ページ設定の追加
# -------------------------
def set_page_break(ws,break_row):
  # 既存クリア
  ws.row_breaks = RowBreak()
  # 改ページ開始行を計算
  start_break_row = 1 + break_row
  # 改ページ追加
  for i in range(start_break_row, ws.max_row, break_row):
      ws.row_breaks.append(Break(id=i))
  # 印刷設定
  ws.page_setup.fitToHeight = False
  ws.page_setup.fitToWidth = False