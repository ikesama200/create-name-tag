from gui.main_window import MainWindow
from utils.logger import setup_logger

if __name__ == "__main__":
    setup_logger()
    app = MainWindow()
    app.run()

'''
from tkinter import filedialog
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws["A1"] = "Hello"
ws["B1"] = "Excel"

wb.save("sample.xlsx")
'''